import os
import pickle
import sys
from subprocess import call
from openpyxl import load_workbook
from datetime import datetime 
import sys
import sqlite3
import os.path


try:
    from tkinter import *
except ImportError:
    from tkinter import *

try:
    import ttk
    py3 = False
except ImportError:
    import tkinter.ttk as ttk
    py3 = True
details_list=[]


def file_save():
    TOKEN_PRO = details_list[0]
    NAME_PRO = details_list[1]
    ADDRESS_PRO = details_list[2]
    MOBILE_NO_PRO = details_list[3]
    ROOM_NO_PRO = details_list[4]
    PEOPLE_PRO = details_list[5]
    ROOM_TYPE_PRO = details_list[6]
    DAYS_PRO = details_list[7]
    PRICE_PRO = details_list[8]
    
    
    f = open("hotel.dat", "ab")
    a=save(TOKEN_PRO,NAME_PRO,ADDRESS_PRO,MOBILE_NO_PRO,ROOM_NO_PRO,PEOPLE_PRO,ROOM_TYPE_PRO,DAYS_PRO,PRICE_PRO)
    pickle.dump(a,f,protocol=2)
    f.close()

    now = datetime.now()
    IN_TIME = str(now.strftime("%d-%b-%Y  %H:%M:%S"))
    IN_TIME_PRO = "IN - " + IN_TIME
    listr=[str(TOKEN_PRO),str(NAME_PRO),str(ADDRESS_PRO),str(MOBILE_NO_PRO),str(ROOM_NO_PRO),str(PEOPLE_PRO),str(ROOM_TYPE_PRO),str(DAYS_PRO),str(PRICE_PRO),str(IN_TIME_PRO)]

    myVars = {'A':TOKEN_PRO,'B':NAME_PRO,"C":ADDRESS_PRO,"D":MOBILE_NO_PRO,"E":ROOM_NO_PRO,"F":PEOPLE_PRO,"G":ROOM_TYPE_PRO, "H":DAYS_PRO,"I":PRICE_PRO,"J":IN_TIME_PRO}
    

    myFileName=r'Data Base.xlsx'
    wb = load_workbook(filename=myFileName)
    ws = wb['Sheet']
    newRowLocation = ws.max_row + 1
    lastcolumn = ws.max_column
    
    for i in range(0,lastcolumn):
        ws.cell(column=i+1,row=newRowLocation, value=listr[i])
        wb.save(filename=myFileName)
    wb.close()

    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    
    cur.execute("CREATE TABLE IF NOT EXISTS hotel(Token_No, Name, Address, Mobile_No, Room_No, No_of_people, Room_type, No_of_days, Price, Check_IN_OUT_time)")

    cur.execute("INSERT INTO hotel (Token_No, Name, Address, Mobile_No, Room_No, No_of_people, Room_type, No_of_days, Price, Check_IN_OUT_time) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",(TOKEN_PRO, NAME_PRO, ADDRESS_PRO, MOBILE_NO_PRO, ROOM_NO_PRO, PEOPLE_PRO, ROOM_TYPE_PRO, DAYS_PRO, PRICE_PRO, IN_TIME_PRO))
    conn.commit()

    for row in cur.execute("SELECT Token_No, Name, Address, Mobile_No, Room_No, No_of_people, Room_type, No_of_days, Price, Check_IN_OUT_time FROM hotel"):
        print(row)
    conn.close()
    
    fo=open("recipt.txt","w+")
    for h in range(0,9):
        fo.write(listr[h]+"\r\n")
    fo.close()
    call(["python","token_and_room_no.py"])
    os.remove("recipt.txt")
    restart_program()






u = list()
Delux = (1, 2, 3, 4, 5, 6, 7, 8, 9, 10)
Full_Delux = (11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25)
General = (26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45)
Joint_Room = (46, 47, 48, 49, 50, 46, 47, 48, 49, 50)
tok = ("#000001","#000002","#000003","#000004","#000005","#000006","#000007","#000008","#000009","#0000010","#000011","#000012","#000013","#000014","#000015","#000016","#000017","#000018","#000019","#000020","#000021","#000022","#000023","#000024","#000025","#000026","#000027","#000028","#000029","#000030","#000031","#000032","#000033","#000034","#000035","#000036","#000037","#000038","#000039","#000040","#000041","#000042","#000043","#000044","#000045","#000046","#000047","#000048","#000049","#000050","#000051","#000052","#000053","#000054","#000055","#000056","#000057","#000058","#000059","#000060","#000061","#000062","#000063","#000064","#000065","#000066","#000067","#000068","#000069","#000070","#000071","#000072","#000073","#000074","#000075","#000076","#000077","#000078","#000079","#000080","#000081","#000082","#000083","#000084","#000085","#000086","#000087","#000088","#000089","#000090","#000091","#000092","#000093","#000094","#000095","#000096","#000097","#000098","#000099","#000100")
m = [9]
G = []
T = []
def restart_program():
    """Restarts the current program.
    Note: this function does not return. Any cleanup action (like
    saving data) must be done before calling this function."""
    python = sys.executable
    os.execl(python, python, * sys.argv)


class save:
    def __init__(self,TOKEN_PRO,NAME_PRO,ADDRESS_PRO,MOBILE_NO_PRO,ROOM_NO_PRO,PEOPLE_PRO,ROOM_TYPE_PRO,DAYS_PRO,PRICE_PRO):
        self.token=TOKEN_PRO
        self.name=NAME_PRO
        self.address=ADDRESS_PRO
        self.mobile_no=MOBILE_NO_PRO
        self.room_no=ROOM_NO_PRO
        self.no_of_people=PEOPLE_PRO
        self.room_type=ROOM_TYPE_PRO
        self.no_of_days=DAYS_PRO
        self.price=PRICE_PRO

        print(self.token,self.name,self.address,self.mobile_no,self.room_no,self.no_of_people,self.room_type,self.no_of_days,self.price)


class HOTEL_MANGMENT_checkin:


    def __init__(self):
        self.token=""
        self.NAME=""
        self.ADDERESS=""
        self.MOBILE=""
        self.PEOPLE=0
        self.ROOM_TYPE=""
        self.DAYS=0
        self.price=0
        self.room=0

        root = Tk()

        def chk_name(pointless=None):
            Name = 0; add = 0; mo = 0; d = 0; p = 0
            
            while True:

                self.k = str(self.name.get())
                a = self.k.isdigit()
                if len(self.k) != 0 and a != True:
                    self.NAME=self.k
                    self.Text1.insert(INSERT, "name has been saved""\n")
                    Name = 1
                    break
                else:
                    self.Text1.insert(INSERT, "invalid input please enter the valid name""\n")
                    Name = 0
                    break

            while True:

                self.g = str(self.addr.get())
                ak = self.g.isdigit()
                if len(self.g)!= 0 and ak!=True:
                    self.ADDERESS=self.g
                    self.Text1.insert(INSERT, "address has been saved""\n")
                    add = 1
                    break
                else:
                    self.Text1.insert(INSERT, "invalid input please enter the valid address""\n")
                    add = 0
                    break

            while True:

                self.h = str(self.mobile.get())
                if self.h.isdigit() == True and len(self.h) != 0 and len(self.h) == 10:
                    self.MOBILE = self.h
                    self.Text1.insert(INSERT, "mobile number has been saved""\n")
                    mo = 1
                    break
                else:
                    self.Text1.insert(INSERT, "invalid input please enter the valid mobile number""\n")
                    mo = 0
                    break

            while True:

                self.l = str(self.days.get())
                if self.l.isdigit() == True and len(self.l) != 0 and self.l != 0:
                    self.DAYS = int(self.l)
                    self.Text1.insert(INSERT, "days has been saved""\n")
                    d = 1
                    break
                else:
                    self.Text1.insert(INSERT, "invalid input ""\n")
                    d = 0
                    break

            while True:

                self.o = str(self.people.get())
                if self.o.isdigit() == True and len(self.o) != 0 and self.o != 0:
                    self.PEOPLE = self.o
                    self.Text1.insert(INSERT, "people has been saved""\n")
                    p = 1
                    break
                else:
                    self.Text1.insert(INSERT, "invalid input""\n")
                    p = 0
                    break

            if Name == 0 or add == 0 or mo == 0 or d == 0 or p == 0:
                self.Button4['state'] = DISABLED
            else:
                self.Button4['state'] = NORMAL
                root.bind('<Return>', submit_clicked)

        def enter(self):
            self.name = self.NAME
            self.address = self.ADDERESS
            self.mobile_no = self.MOBILE
            self.no_of_people = int(self.PEOPLE)
            self.no_of_days = int(self.DAYS)
        
        def bill(self):
            price = 0
            if self.ch == 1 and self.p == 2:
                self.ROOM_TYPE = "Delux"
                price = self.price + (1500 * self.no_of_days * self.no_of_people)
                self.price = price - (price * 0.1)
                m[0] = 1
            elif self.ch == 1 and self.p == 1:
                self.ROOM_TYPE = "Delux"
                self.price = self.price + (1500 * self.no_of_days * self.no_of_people)
                m[0] = 1
            elif self.ch == 2 and self.p == 2:
                self.ROOM_TYPE = "Full Delux"
                price = self.price + (1700 * self.no_of_days * self.no_of_people)
                self.price = price - (price * 0.1)
                m[0] = 2
            elif self.ch == 2 and self.p == 1:
                self.ROOM_TYPE = "Full Delux"
                self.price = self.price + (1700 * self.no_of_days * self.no_of_people)
                m[0] = 2
            elif self.ch == 3 and self.p == 2:
                self.ROOM_TYPE = "General"
                price = self.price + (1000 * self.no_of_days * self.no_of_people)
                self.price = price - (price * 0.1)
                m[0] = 3
            elif self.ch == 3 and self.p == 1:
                self.ROOM_TYPE = "General"
                self.price = self.price + (1000 * self.no_of_days * self.no_of_people)
                m[0] = 3
            elif self.ch == 4 and self.p == 2:
                self.ROOM_TYPE = "Joint"
                price = self.price + (700 * self.no_of_days * self.no_of_people)
                self.price = price - (price * 0.1)
                m[0] = 4
            elif self.ch == 4 and self.p == 1:
                self.ROOM_TYPE = "Joint"
                self.price = self.price + (700 * self.no_of_days * self.no_of_people)
                m[0] = 4
            
            self.room_type = self.ROOM_TYPE
            
        def token(self):
            T = []
            myfile = r'Data Base.xlsx'
            wb = load_workbook(filename=myfile)
            ws = wb['Sheet']
            row = ws.max_row+1
            for i in range(2,row):
                CELL = ws.cell(row=i,column=1)
                k = CELL.value
                T.append(k)
                continue
            for t in tok:
                if t not in T:
                    self.token = t
                    break
                else:
                    continue
            self.token = t
            wb.close()

        
        def room_no(self):

            if m[0] == 1:
                a = Delux
            elif m[0] == 2:
                a = Full_Delux
            elif m[0] == 3:
                a = General
            elif m[0] == 4:
                a = Joint_Room

            G = []
            f2 = open("hotel.dat", "rb")
            try:
                while True:
                    s = pickle.load(f2)

                    k = s.room_no
                    G.append(k)
                    continue

            except EOFError:
                pass
        
            for r in a:
                if r not in G:
                    self.room = r
                    break
                else:
                    continue
            self.room = r
            f2.close()

            details_list.clear()
            
            details_list.append(self.token)
            details_list.append(self.name)
            details_list.append(self.address)
            details_list.append(self.mobile_no)
            details_list.append(self.room)
            details_list.append(self.no_of_people)
            details_list.append(self.room_type)
            details_list.append(self.no_of_days)
            details_list.append(self.price)

            file_save()



        def submit_clicked(pointless=None):
            if self.var1.get()==1 and self.var2.get()==0 and self.var3.get()==0 and self.var4.get()==0 and self.var5.get()==0 and self.var6.get()==1:
                self.ch=1
                self.p=1
                token(self)
                enter(self)
                bill(self)
                room_no(self)
            elif self.var1.get() == 1 and self.var2.get() == 0 and self.var3.get() == 0 and self.var4.get() == 0 and self.var5.get() == 1 and self.var6.get() == 0:
                self.ch = 1
                self.p = 2
                token(self)
                enter(self)
                bill(self)
                room_no(self)
            elif self.var1.get() == 0 and self.var2.get() == 1 and self.var3.get() == 0 and self.var4.get() == 0 and self.var5.get() == 0 and self.var6.get() == 1:
                self.ch = 2
                self.p = 1
                token(self)
                enter(self)
                bill(self)
                room_no(self)
            elif self.var1.get() == 0 and self.var2.get() == 1 and self.var3.get() == 0 and self.var4.get() == 0 and self.var5.get() == 1 and self.var6.get() ==0 :
                self.ch = 2
                self.p = 2
                token(self)
                enter(self)
                bill(self)
                room_no(self)
            elif self.var1.get() == 0 and self.var2.get() == 0 and self.var3.get() == 1 and self.var4.get() == 0 and self.var5.get() == 0 and self.var6.get() == 1:
                self.ch = 3
                self.p = 1
                token(self)
                enter(self)
                bill(self)
                room_no(self)
            elif self.var1.get() == 0 and self.var2.get() == 0 and self.var3.get() == 1 and self.var4.get() == 0 and self.var5.get() == 1 and self.var6.get() == 0:
                self.ch = 3
                self.p = 2
                token(self)
                enter(self)
                bill(self)
                room_no(self)
            elif self.var1.get() == 0 and self.var2.get() == 0 and self.var3.get() == 0 and self.var4.get() == 1 and self.var5.get() == 0 and self.var6.get() == 1:
                self.ch = 4
                self.p = 1
                token(self)
                enter(self)
                bill(self)
                room_no(self)
            elif self.var1.get() == 0 and self.var2.get() == 0 and self.var3.get() == 0 and self.var4.get() == 1 and self.var5.get() == 1 and self.var6.get() == 0:
                self.ch = 4
                self.p = 2
                token(self)
                enter(self)
                bill(self)
                room_no(self)
            else:
                self.Text1.insert(INSERT, "invalid choice please input a valid choice""\n")

        '''This class configures and populates the toplevel window.
           top is the toplevel containing window.'''
        _bgcolor = '#483D8B'
        _fgcolor = '#000000'
        _compcolor = '#ffffff'
        _ana1color = '#ffffff'
        _ana2color = '#ffffff'
        font10 = "-family {Helvetica Neue Light} -size 15 -weight normal -slant"  \
            " roman -underline 0 -overstrike 0"
        font11 = "-family {Segoe UI} -size 30 -weight bold -slant "  \
            "roman -underline 0 -overstrike 0"
        font12 = "-family {Segoe UI} -size 18 -weight bold -slant "  \
            "roman -underline 0 -overstrike 0"
        font13 = "-family {Segoe UI} -size 17 -weight bold -slant "  \
            "roman -underline 0 -overstrike 0"
        font14 = "-family {Segoe UI} -size 16 -weight bold -slant "  \
            "roman -underline 0 -overstrike 0"
        font15 = "-family {Segoe UI} -size 19 -weight bold -slant "  \
            "roman -underline 0 -overstrike 0"
        font16 = "-family {Segoe UI} -size 15 -weight bold -slant "  \
            "roman -underline 0 -overstrike 0"
        font17 = "-family {Segoe UI} -size 11 -weight bold -slant"  \
            " roman -underline 0 -overstrike 0"
        font18 = "-family {Segoe UI} -size 18 -weight bold -slant"  \
            " roman -underline 0 -overstrike 0"
        font9 = "-family {Segoe UI} -size 9 -weight normal -slant "  \
            "roman -underline 0 -overstrike 0"

        root.geometry("1069x742")
        root.title("HOTEL MANGMENT")
        root.configure(background="#483D8B")
        root.configure(highlightbackground="#ffffff")
        root.configure(highlightcolor="black")

        self.Text1 = Text(root)
        self.Text1.place(relx=0.03,rely=0.789, relheight=0.2, relwidth=0.93)
        self.Text1.configure(background="#A2CD5A")
        self.Text1.configure(font=font10)
        self.Text1.configure(foreground="black")
        self.Text1.configure(highlightbackground="#ffffff")
        self.Text1.configure(highlightcolor="black")
        self.Text1.configure(insertbackground="black")
        self.Text1.configure(selectbackground="#e6e6e6")
        self.Text1.configure(selectforeground="black")
        self.Text1.configure(width=994)
        self.Text1.configure(wrap=WORD)

        self.Frame1 = Frame(root)
        self.Frame1.place(relx=0.03, rely=0.05, relheight=0.12, relwidth=0.93)
        self.Frame1.configure(relief=GROOVE)
        self.Frame1.configure(borderwidth="2")
        self.Frame1.configure(relief=GROOVE)
        self.Frame1.configure(background="#B4EEB4")
        self.Frame1.configure(highlightbackground="#ffffff")
        self.Frame1.configure(highlightcolor="black")
        self.Frame1.configure(width=995)

        self.Message1 = Message(self.Frame1)
        self.Message1.place(relx=0.04, rely=0.11, relheight=0.84, relwidth=0.5)
        self.Message1.configure(background="#B4EEB4")
        self.Message1.configure(font=font11)
        self.Message1.configure(foreground="#000000")
        self.Message1.configure(highlightbackground="#ffffff")
        self.Message1.configure(highlightcolor="black")
        self.Message1.configure(text='''YOU CLICKED ON''')
        self.Message1.configure(width=496)

        self.Message2 = Message(self.Frame1)
        self.Message2.place(relx=0.52, rely=0.19, relheight=0.74, relwidth=0.07)
        self.Message2.configure(background="#B4EEB4")
        self.Message2.configure(font=font11)
        self.Message2.configure(foreground="#000000")
        self.Message2.configure(highlightbackground="#ffffff")
        self.Message2.configure(highlightcolor="black")
        self.Message2.configure(text=''':''')
        self.Message2.configure(width=66)

        self.Message3 = Message(self.Frame1)
        self.Message3.place(relx=0.57, rely=0.11, relheight=0.79, relwidth=0.35)
        self.Message3.configure(background="#B4EEB4")
        self.Message3.configure(font=font11)
        self.Message3.configure(foreground="#000000")
        self.Message3.configure(highlightbackground="#ffffff")
        self.Message3.configure(highlightcolor="black")
        self.Message3.configure(text='''CHECK INN''')
        self.Message3.configure(width=347)

        self.menubar = Menu(root,font=font9,bg=_bgcolor,fg=_fgcolor)
        root.configure(menu = self.menubar)

        self.Frame2 = Frame(root)
        self.Frame2.place(relx=0.03, rely=0.18, relheight=0.6, relwidth=0.93)
        self.Frame2.configure(relief=GROOVE)
        self.Frame2.configure(borderwidth="2")
        self.Frame2.configure(relief=GROOVE)
        self.Frame2.configure(background="#BCEE68")
        self.Frame2.configure(highlightbackground="#ffffff")
        self.Frame2.configure(highlightcolor="black")
        self.Frame2.configure(width=995)

        self.Label3 = Label(self.Frame2)
        self.Label3.place(relx=0.05, rely=0.01, height=47, width=289)
        self.Label3.configure(activebackground="#ffffff")
        self.Label3.configure(activeforeground="black")
        self.Label3.configure(background="#BCEE68")
        self.Label3.configure(disabledforeground="#bfbfbf")
        self.Label3.configure(font=font12)
        self.Label3.configure(foreground="#000000")
        self.Label3.configure(highlightbackground="#ffffff")
        self.Label3.configure(highlightcolor="black")
        self.Label3.configure(text='''ENTER YOUR NAME''')

        self.Message4 = Message(self.Frame2)
        self.Message4.place(relx=0.41, rely=0.01, relheight=0.1, relwidth=0.05)
        self.Message4.configure(background="#BCEE68")
        self.Message4.configure(font=font13)
        self.Message4.configure(foreground="#000000")
        self.Message4.configure(highlightbackground="#ffffff")
        self.Message4.configure(highlightcolor="black")
        self.Message4.configure(text=''':''')
        self.Message4.configure(width=46)

        self.Entry3 = Entry(self.Frame2)
        self.name=StringVar()
        self.Entry3.place(relx=0.47, rely=0.019,height=34, relwidth=0.43)
        self.Entry3.configure(background="white")
        self.Entry3.configure(disabledforeground="#bfbfbf")
        self.Entry3.configure(font=font10)
        self.Entry3.configure(foreground="#000000")
        self.Entry3.configure(highlightbackground="#ffffff")
        self.Entry3.configure(highlightcolor="black")
        self.Entry3.configure(insertbackground="black")
        self.Entry3.configure(selectbackground="#e6e6e6")
        self.Entry3.configure(selectforeground="black")
        self.Entry3.configure(textvariable=self.name)

        self.Label5 = Label(self.Frame2)
        self.Label5.place(relx=0.045, rely=0.107, height=47, width=334)
        self.Label5.configure(activebackground="#ffffff")
        self.Label5.configure(activeforeground="black")
        self.Label5.configure(background="#BCEE68")
        self.Label5.configure(disabledforeground="#bfbfbf")
        self.Label5.configure(font=font12)
        self.Label5.configure(foreground="#000000")
        self.Label5.configure(highlightbackground="#ffffff")
        self.Label5.configure(highlightcolor="black")
        self.Label5.configure(text='''ENTER YOUR ADDRESS''')

        self.Message6 = Message(self.Frame2)
        self.Message6.place(relx=0.415, rely=0.107, relheight=0.09, relwidth=0.04)
        self.Message6.configure(background="#BCEE68")
        self.Message6.configure(font=font13)
        self.Message6.configure(foreground="#000000")
        self.Message6.configure(highlightbackground="#ffffff")
        self.Message6.configure(highlightcolor="black")
        self.Message6.configure(text=''':''')
        self.Message6.configure(width=36)

        self.Entry5 = Entry(self.Frame2)
        self.addr = StringVar()
        self.Entry5.place(relx=0.47, rely=0.12,height=34, relwidth=0.43)
        self.Entry5.configure(background="white")
        self.Entry5.configure(disabledforeground="#bfbfbf")
        self.Entry5.configure(font=font10)
        self.Entry5.configure(foreground="#000000")
        self.Entry5.configure(highlightbackground="#ffffff")
        self.Entry5.configure(highlightcolor="black")
        self.Entry5.configure(insertbackground="black")
        self.Entry5.configure(selectbackground="#e6e6e6")
        self.Entry5.configure(selectforeground="black")
        self.Entry5.configure(textvariable=self.addr)

        self.Label4 = Label(self.Frame2)
        self.Label4.place(relx=0.045, rely=0.209, height=47, width=329)
        self.Label4.configure(activebackground="#ffffff")
        self.Label4.configure(activeforeground="black")
        self.Label4.configure(background="#BCEE68")
        self.Label4.configure(disabledforeground="#bfbfbf")
        self.Label4.configure(font=font12)
        self.Label4.configure(foreground="#000000")
        self.Label4.configure(highlightbackground="#ffffff")
        self.Label4.configure(highlightcolor="black")
        self.Label4.configure(text='''ENTER YOUR NUMBER''')

        self.Message8 = Message(self.Frame2)
        self.Message8.place(relx=0.42, rely=0.2, relheight=0.11, relwidth=0.03)
        self.Message8.configure(background="#BCEE68")
        self.Message8.configure(font=font13)
        self.Message8.configure(foreground="#000000")
        self.Message8.configure(highlightbackground="#ffffff")
        self.Message8.configure(highlightcolor="black")
        self.Message8.configure(text=''':''')
        self.Message8.configure(width=26)

        self.Entry4 = Entry(self.Frame2)
        self.mobile=StringVar()
        self.Entry4.place(relx=0.47, rely=0.22,height=34, relwidth=0.43)
        self.Entry4.configure(background="white")
        self.Entry4.configure(disabledforeground="#bfbfbf")
        self.Entry4.configure(font=font10)
        self.Entry4.configure(foreground="#000000")
        self.Entry4.configure(highlightbackground="#ffffff")
        self.Entry4.configure(highlightcolor="black")
        self.Entry4.configure(insertbackground="black")
        self.Entry4.configure(selectbackground="#e6e6e6")
        self.Entry4.configure(selectforeground="black")
        self.Entry4.configure(textvariable=self.mobile)

        self.Label1 = Label(self.Frame2)
        self.Label1.place(relx=0.05, rely=0.315, height=44, width=260)
        self.Label1.configure(background="#BCEE68")
        self.Label1.configure(disabledforeground="#bfbfbf")
        self.Label1.configure(font=font13)
        self.Label1.configure(foreground="#000000")
        self.Label1.configure(text='''NUMBER OF DAYS''')

        self.Message8 = Message(self.Frame2)
        self.Message8.place(relx=0.42, rely=0.315, relheight=0.11, relwidth=0.03)
        self.Message8.configure(background="#BCEE68")
        self.Message8.configure(font=font13)
        self.Message8.configure(foreground="#000000")
        self.Message8.configure(highlightbackground="#ffffff")
        self.Message8.configure(highlightcolor="black")
        self.Message8.configure(text=''':''')
        self.Message8.configure(width=26)

        self.Entry1 = Entry(self.Frame2)
        self.days=StringVar()
        self.Entry1.place(relx=0.47, rely=0.325, height=34, relwidth=0.43)
        self.Entry1.configure(background="white")
        self.Entry1.configure(disabledforeground="#bfbfbf")
        self.Entry1.configure(font=font10)
        self.Entry1.configure(foreground="#000000")
        self.Entry1.configure(insertbackground="black")
        self.Entry1.configure(width=424)
        self.Entry1.configure(textvariable=self.days)

        self.Label8 = Label(self.Frame2)
        self.Label8.place(relx=0.04, rely=0.42, height=44, width=260)
        self.Label8.configure(activebackground="#ffffff")
        self.Label8.configure(activeforeground="black")
        self.Label8.configure(background="#BCEE68")
        self.Label8.configure(disabledforeground="#bfbfbf")
        self.Label8.configure(font=font13)
        self.Label8.configure(foreground="#000000")
        self.Label8.configure(highlightbackground="#ffffff")
        self.Label8.configure(highlightcolor="black")
        self.Label8.configure(text='''NO OF PEOPLE''')

        self.Message8 = Message(self.Frame2)
        self.Message8.place(relx=0.42, rely=0.42, relheight=0.11, relwidth=0.03)
        self.Message8.configure(background="#BCEE68")
        self.Message8.configure(font=font13)
        self.Message8.configure(foreground="#000000")
        self.Message8.configure(highlightbackground="#ffffff")
        self.Message8.configure(highlightcolor="black")
        self.Message8.configure(text=''':''')
        self.Message8.configure(width=26)

        self.Entry2 = Entry(self.Frame2)
        self.people=StringVar()
        self.Entry2.place(relx=0.47, rely=0.42, height=34, relwidth=0.43)
        self.Entry2.configure(background="white")
        self.Entry2.configure(disabledforeground="#bfbfbf")
        self.Entry2.configure(font=font10)
        self.Entry2.configure(foreground="#000000")
        self.Entry2.configure(insertbackground="black")
        self.Entry2.configure(width=424)
        self.Entry2.configure(textvariable=self.people)

        self.Label6 = Label(self.Frame2)
        self.Label6.place(relx=0.32, rely=0.53, height=48, width=296)
        self.Label6.configure(activebackground="#ffffff")
        self.Label6.configure(activeforeground="black")
        self.Label6.configure(background="#BCEE68")
        self.Label6.configure(disabledforeground="#bfbfbf")
        self.Label6.configure(font=font13)
        self.Label6.configure(foreground="#000000")
        self.Label6.configure(highlightbackground="#ffffff")
        self.Label6.configure(highlightcolor="black")
        self.Label6.configure(text='''CHOOSE YOUR ROOM''')

        self.Label7 = Label(self.Frame2)
        self.Label7.place(relx=0.3, rely=0.79, height=48, width=300)
        self.Label7.configure(activebackground="#ffffff")
        self.Label7.configure(activeforeground="black")
        self.Label7.configure(background="#BCEE68")
        self.Label7.configure(disabledforeground="#bfbfbf")
        self.Label7.configure(font=font14)
        self.Label7.configure(foreground="#000000")
        self.Label7.configure(highlightbackground="#ffffff")
        self.Label7.configure(highlightcolor="black")
        self.Label7.configure(text='''CHOOSE PAYMENT METHOD''')

        self.Checkbutton1 = Checkbutton(self.Frame2)
        self.var1 = IntVar()
        self.Checkbutton1.place(relx=0.15, rely=0.58, relheight=0.17, relwidth=0.14)
        self.Checkbutton1.configure(activebackground="#ffffff")
        self.Checkbutton1.configure(activeforeground="#000000")
        self.Checkbutton1.configure(background="#BCEE68")
        self.Checkbutton1.configure(disabledforeground="#bfbfbf")
        self.Checkbutton1.configure(font=font14)
        self.Checkbutton1.configure(foreground="#000000")
        self.Checkbutton1.configure(highlightbackground="#ffffff")
        self.Checkbutton1.configure(highlightcolor="black")
        self.Checkbutton1.configure(justify=LEFT)
        self.Checkbutton1.configure(text='''DELUXE''')
        self.Checkbutton1.configure(variable=self.var1)

        self.Checkbutton2 = Checkbutton(self.Frame2)
        self.var2 = IntVar()
        self.Checkbutton2.place(relx=0.15, rely=0.72, relheight=0.11, relwidth=0.21)
        self.Checkbutton2.configure(activebackground="#ffffff")
        self.Checkbutton2.configure(activeforeground="#000000")
        self.Checkbutton2.configure(background="#BCEE68")
        self.Checkbutton2.configure(disabledforeground="#bfbfbf")
        self.Checkbutton2.configure(font=font13)
        self.Checkbutton2.configure(foreground="#000000")
        self.Checkbutton2.configure(highlightbackground="#ffffff")
        self.Checkbutton2.configure(highlightcolor="black")
        self.Checkbutton2.configure(justify=LEFT)
        self.Checkbutton2.configure(text='''FULL DELUXE''')
        self.Checkbutton2.configure(variable=self.var2)

        self.Checkbutton3 = Checkbutton(self.Frame2)
        self.var3 = IntVar()
        self.Checkbutton3.place(relx=0.5, rely=0.628, relheight=0.09, relwidth=0.16)
        self.Checkbutton3.configure(activebackground="#ffffff")
        self.Checkbutton3.configure(activeforeground="#000000")
        self.Checkbutton3.configure(background="#BCEE68")
        self.Checkbutton3.configure(disabledforeground="#bfbfbf")
        self.Checkbutton3.configure(font=font13)
        self.Checkbutton3.configure(foreground="#000000")
        self.Checkbutton3.configure(highlightbackground="#ffffff")
        self.Checkbutton3.configure(highlightcolor="black")
        self.Checkbutton3.configure(justify=LEFT)
        self.Checkbutton3.configure(text='''GENERAL''')
        self.Checkbutton3.configure(variable=self.var3)

        self.Checkbutton4 = Checkbutton(self.Frame2)
        self.var4 = IntVar()
        self.Checkbutton4.place(relx=0.5, rely=0.71, relheight=0.11, relwidth=0.12)
        self.Checkbutton4.configure(activebackground="#ffffff")
        self.Checkbutton4.configure(activeforeground="#000000")
        self.Checkbutton4.configure(background="#BCEE68")
        self.Checkbutton4.configure(disabledforeground="#bfbfbf")
        self.Checkbutton4.configure(font=font13)
        self.Checkbutton4.configure(foreground="#000000")
        self.Checkbutton4.configure(highlightbackground="#ffffff")
        self.Checkbutton4.configure(highlightcolor="black")
        self.Checkbutton4.configure(justify=LEFT)
        self.Checkbutton4.configure(text='''JOINT''')
        self.Checkbutton4.configure(variable=self.var4)

        self.Checkbutton6 = Checkbutton(self.Frame2)
        self.var6 = IntVar()
        self.Checkbutton6.place(relx=0.153, rely=0.89, relheight=0.1, relwidth=0.15)
        self.Checkbutton6.configure(activebackground="#ffffff")
        self.Checkbutton6.configure(activeforeground="#000000")
        self.Checkbutton6.configure(background="#BCEE68")
        self.Checkbutton6.configure(disabledforeground="#bfbfbf")
        self.Checkbutton6.configure(font=font16)
        self.Checkbutton6.configure(foreground="#000000")
        self.Checkbutton6.configure(highlightbackground="#ffffff")
        self.Checkbutton6.configure(highlightcolor="black")
        self.Checkbutton6.configure(justify=LEFT)
        self.Checkbutton6.configure(text='''By cash''')
        self.Checkbutton6.configure(variable=self.var6)

        self.Checkbutton5 = Checkbutton(self.Frame2)
        self.var5 = IntVar()
        self.Checkbutton5.place(relx=0.485, rely=0.89, relheight=0.1, relwidth=0.3)
        self.Checkbutton5.configure(activebackground="#ffffff")
        self.Checkbutton5.configure(activeforeground="#000000")
        self.Checkbutton5.configure(background="#BCEE68")
        self.Checkbutton5.configure(disabledforeground="#bfbfbf")
        self.Checkbutton5.configure(font=font16)
        self.Checkbutton5.configure(foreground="#000000")
        self.Checkbutton5.configure(highlightbackground="#ffffff")
        self.Checkbutton5.configure(highlightcolor="black")
        self.Checkbutton5.configure(justify=LEFT)
        self.Checkbutton5.configure(text='''By credit/debit card''')
        self.Checkbutton5.configure(variable=self.var5)

        '''self.Message7 = Message(self.Frame2)
        self.Message7.place(relx=0.28, rely=0.46, relheight=0.11, relwidth=0.04)
        self.Message7.configure(background="#BCEE68")
        self.Message7.configure(font=font15)
        self.Message7.configure(foreground="#000000")
        self.Message7.configure(highlightbackground="#ffffff")
        self.Message7.configure(highlightcolor="black")
        self.Message7.configure(text='''#-''')
        #self.Message7.configure(width=41)
                
    
        self.Button1 = Button(self.Frame2)
        self.Button1.place(relx=0.81, rely=0.55, height=33, width=43)
        self.Button1.configure(relief="raised")
        self.Button1.configure(borderwidth=2)
        self.Button1.configure(activebackground="#000000")
        self.Button1.configure(activeforeground="#ffffff")
        self.Button1.configure(background="#ffffff")
        self.Button1.configure(disabledforeground="#bfbfbf")
        self.Button1.configure(foreground="#000000")
        self.Button1.configure(highlightbackground="#ffffff")
        self.Button1.configure(highlightcolor="black")
        self.Button1.configure(font=font17)
        self.Button1.configure(pady="0")
        self.Button1.configure(text='''SAVE''')
        self.Button1.configure(command=chk_name)
        root.bind('<Return>', chk_name)

        self.Button4 = Button(self.Frame2)
        self.Button4.place(relx=0.76, rely=0.66, height=83, width=156)
        self.Button4.configure(relief="raised")
        self.Button4.configure(borderwidth=8)
        self.Button4.configure(activebackground="#000000")
        self.Button4.configure(activeforeground="#ffffff")
        self.Button4.configure(background="#ffffff")
        self.Button4.configure(disabledforeground="#bfbfbf")
        self.Button4.configure(font=font18)
        self.Button4.configure(foreground="#000000")
        self.Button4.configure(highlightbackground="#ffffff")
        self.Button4.configure(highlightcolor="black")
        self.Button4.configure(pady="0")
        self.Button4.configure(state=NORMAL)
        self.Button4.configure(text='''SUBMIT''')
        self.Button4.configure(command=submit_clicked)

        

        root.mainloop()


if __name__ == '__main__':
    hotel=HOTEL_MANGMENT_checkin()